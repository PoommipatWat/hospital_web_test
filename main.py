import os
import io
import csv
import cv2
import base64
import threading
import time
from datetime import datetime, timedelta
from flask import Flask, render_template, request, redirect, url_for, flash, send_file, Response, jsonify
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import desc
import random # เพื่อสร้างข้อมูลตัวอย่าง
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from flask_socketio import SocketIO
from sqlalchemy import exc as sqlalchemy_exc

from robot_controller import *

# สร้าง Flask app
app = Flask(__name__)
app.config['SECRET_KEY'] = 'your-secret-key'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///vitals.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

socketio = SocketIO(app, cors_allowed_origins="*")

# เริ่มต้น ROS2 และ RobotController
rclpy.init()
robot_controller = RobotController(socketio_instance=socketio)

# รัน ROS2 spin ใน thread แยก
ros_thread = threading.Thread(target=spin_robot_controller, args=(robot_controller,), daemon=True)
ros_thread.start()


# ตัวแปรกลอบอลสำหรับเก็บกระแสวิดีโอล่าสุด
global_frame = None
camera_active = False
camera_thread = None
lock = threading.Lock()

# สร้างโฟลเดอร์สำหรับเก็บรูปภาพ (ถ้ายังไม่มี)
UPLOAD_FOLDER = 'static/uploads'
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

# สร้าง database instance
db = SQLAlchemy(app)

# สร้างโมเดลสำหรับเก็บข้อมูล vital signs
class VitalSign(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    patient_id = db.Column(db.String(20), nullable=False)
    patient_name = db.Column(db.String(100), nullable=False)
    room_number = db.Column(db.String(20))  # เพิ่มหมายเลขห้อง
    weight = db.Column(db.Float)  # เพิ่มน้ำหนัก (กิโลกรัม)
    height = db.Column(db.Float)  # เพิ่มส่วนสูง (เซนติเมตร)
    timestamp = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    temperature = db.Column(db.Float)
    heart_rate = db.Column(db.Integer)
    blood_pressure_sys = db.Column(db.Integer)
    blood_pressure_dia = db.Column(db.Integer)
    respiratory_rate = db.Column(db.Integer)
    oxygen_saturation = db.Column(db.Integer)
    
    def __repr__(self):
        return f'<VitalSign {self.patient_id} {self.timestamp}>'

# สร้างฐานข้อมูลและตาราง
with app.app_context():
    db.create_all()

# สร้างข้อมูลตัวอย่าง
def create_sample_data():
    # ลบข้อมูลเก่า
    db.session.query(VitalSign).delete()
    
    patients = [
        {"id": "P001", "name": "สมชาย ใจดี", "room": "101A", "weight": 65.5, "height": 170.5},
        {"id": "P002", "name": "สมศรี รักเรียน", "room": "102B", "weight": 55.0, "height": 160.0},
        {"id": "P003", "name": "มานี มานะ", "room": "103C", "weight": 70.2, "height": 175.0}
    ]
    
    # สร้างข้อมูลย้อนหลัง 3 วัน โดยวัดทุก 6 ชั่วโมง
    # ใช้เวลาประเทศไทย
    thai_now = datetime.now() + timedelta(hours=7)
    
    for patient in patients:
        for days_ago in range(3):
            for hours in [0, 6, 12, 18]:
                timestamp = thai_now - timedelta(days=days_ago, hours=thai_now.hour - hours)
                
                vital = VitalSign(
                    patient_id=patient["id"],
                    patient_name=patient["name"],
                    room_number=patient["room"],
                    weight=patient["weight"],
                    height=patient["height"],
                    timestamp=timestamp,
                    temperature=round(random.uniform(36.0, 38.5), 1),
                    heart_rate=random.randint(60, 100),
                    blood_pressure_sys=random.randint(110, 140),
                    blood_pressure_dia=random.randint(60, 90),
                    respiratory_rate=random.randint(12, 20),
                    oxygen_saturation=random.randint(95, 100)
                )
                db.session.add(vital)
    
    db.session.commit()

# สร้างเส้นทางหลัก
@app.route("/")
def index():
    # ดึงข้อมูลผู้ป่วยทั้งหมด (unique) โดยใช้ subquery เพื่อให้แน่ใจว่าไม่มีการซ้ำกัน
    from sqlalchemy import func
    
    # ดึงข้อมูลผู้ป่วยล่าสุดของแต่ละรหัสผู้ป่วย
    subquery = db.session.query(
        VitalSign.patient_id,
        func.max(VitalSign.timestamp).label('latest_timestamp')
    ).group_by(VitalSign.patient_id).subquery()
    
    patients_data = db.session.query(
        VitalSign.patient_id, 
        VitalSign.patient_name,
        VitalSign.room_number,
        VitalSign.weight,
        VitalSign.height
    ).join(
        subquery,
        db.and_(
            VitalSign.patient_id == subquery.c.patient_id,
            VitalSign.timestamp == subquery.c.latest_timestamp
        )
    ).all()
    
    # เพิ่มสถานะของผู้ป่วยแต่ละราย (ปกติ, ผิดปกติ, ไม่มีข้อมูล)
    patients_with_status = []
    
    for patient_record in patients_data:
        patient_id = patient_record[0]
        
        # ดึงข้อมูล vital signs ล่าสุดของผู้ป่วย
        latest_vital = VitalSign.query.filter_by(patient_id=patient_id).filter(
            VitalSign.temperature != None,
            VitalSign.heart_rate != None,
            VitalSign.blood_pressure_sys != None
        ).order_by(desc(VitalSign.timestamp)).first()
        
        # กำหนดสถานะตามข้อมูลล่าสุด
        if latest_vital:
            abnormal_values = check_abnormal_values(latest_vital)
            status = "abnormal" if abnormal_values else "normal"
        else:
            status = "no-data"
        
        # เพิ่มข้อมูลพร้อมสถานะ
        patients_with_status.append(
            (*patient_record, status)
        )
    
    return render_template("index.html", patients=patients_with_status)

# เส้นทางสำหรับเพิ่มผู้ป่วยใหม่
@app.route("/add_patient", methods=["GET", "POST"])
def add_patient():
    if request.method == "POST":
        try:
            patient_id = request.form["patient_id"]
            patient_name = request.form["patient_name"]
            room_number = request.form["room_number"] if "room_number" in request.form and request.form["room_number"] else None
            
            # แปลงค่าน้ำหนักและส่วนสูงเป็นตัวเลข (ถ้ามีการกรอก)
            weight = None
            height = None
            if "weight" in request.form and request.form["weight"]:
                weight = float(request.form["weight"])
            if "height" in request.form and request.form["height"]:
                height = float(request.form["height"])
            
            # ตรวจสอบว่ามีรหัสผู้ป่วยนี้อยู่แล้วหรือไม่
            existing_patient = db.session.query(VitalSign).filter_by(patient_id=patient_id).first()
            if existing_patient:
                flash(f"รหัสผู้ป่วย {patient_id} มีอยู่ในระบบแล้ว", "danger")
                return redirect(url_for("add_patient"))
            
            # เพิ่มข้อมูลผู้ป่วยใหม่โดยไม่มีข้อมูล vital signs
            # ข้อมูลผู้ป่วยจะแสดงในรายการผู้ป่วย แต่ยังไม่มีการบันทึก vital signs
            # สามารถเพิ่มข้อมูล vital signs ได้ในภายหลัง
            flash(f"เพิ่มผู้ป่วย {patient_name} สำเร็จ โปรดบันทึกข้อมูล Vital Signs", "success")
            
            # เพิ่มข้อมูลผู้ป่วยในฐานข้อมูลชั่วคราว - ใช้วิธีนี้เพื่อให้ผู้ป่วยปรากฏในรายการผู้ป่วย
            # แต่ไม่มีข้อมูล vital signs จริง
            # บันทึกเวลาเป็น UTC แต่บวกเวลา 7 ชั่วโมงเพื่อให้เป็นเวลาประเทศไทย
            thai_time = datetime.now() + timedelta(hours=7)
            
            temp_vital = VitalSign(
                patient_id=patient_id,
                patient_name=patient_name,
                room_number=room_number,
                weight=weight,
                height=height,
                timestamp=thai_time,
                temperature=None,
                heart_rate=None,
                blood_pressure_sys=None,
                blood_pressure_dia=None,
                respiratory_rate=None,
                oxygen_saturation=None
            )
            db.session.add(temp_vital)
            db.session.commit()
            
            return redirect(url_for("index"))
        except Exception as e:
            flash(f"เกิดข้อผิดพลาด: {str(e)}", "danger")
    
    return render_template("add_patient.html")

# เส้นทางสำหรับลบผู้ป่วย
@app.route("/delete_patient/<patient_id>", methods=["POST"])
def delete_patient(patient_id):
    try:
        # ลบข้อมูล vital signs ทั้งหมดของผู้ป่วย
        vitals = VitalSign.query.filter_by(patient_id=patient_id).all()
        patient_name = vitals[0].patient_name if vitals else "ไม่ทราบชื่อ"
        
        for vital in vitals:
            db.session.delete(vital)
        
        db.session.commit()
        flash(f"ลบข้อมูลผู้ป่วย {patient_name} (รหัส: {patient_id}) สำเร็จ", "success")
    except Exception as e:
        flash(f"เกิดข้อผิดพลาด: {str(e)}", "danger")
    
    return redirect(url_for("index"))

# ฟังก์ชันประเมินค่าผิดปกติ
def check_abnormal_values(vital):
    abnormal_values = []
    
    # ตรวจสอบอุณหภูมิ
    if vital.temperature is not None:
        if vital.temperature < 36.0:
            abnormal_values.append(f"อุณหภูมิต่ำกว่าปกติ ({vital.temperature} °C)")
        elif vital.temperature > 38.0:
            abnormal_values.append(f"อุณหภูมิสูงกว่าปกติ ({vital.temperature} °C)")
    
    # ตรวจสอบอัตราการเต้นของหัวใจ
    if vital.heart_rate is not None:
        if vital.heart_rate < 60:
            abnormal_values.append(f"อัตราการเต้นของหัวใจต่ำกว่าปกติ ({vital.heart_rate} ครั้ง/นาที)")
        elif vital.heart_rate > 100:
            abnormal_values.append(f"อัตราการเต้นของหัวใจสูงกว่าปกติ ({vital.heart_rate} ครั้ง/นาที)")
    
    # ตรวจสอบความดันโลหิต
    if vital.blood_pressure_sys is not None and vital.blood_pressure_dia is not None:
        if vital.blood_pressure_sys < 90:
            abnormal_values.append(f"ความดันโลหิตตัวบนต่ำกว่าปกติ ({vital.blood_pressure_sys} mmHg)")
        elif vital.blood_pressure_sys > 140:
            abnormal_values.append(f"ความดันโลหิตตัวบนสูงกว่าปกติ ({vital.blood_pressure_sys} mmHg)")
        
        if vital.blood_pressure_dia < 60:
            abnormal_values.append(f"ความดันโลหิตตัวล่างต่ำกว่าปกติ ({vital.blood_pressure_dia} mmHg)")
        elif vital.blood_pressure_dia > 90:
            abnormal_values.append(f"ความดันโลหิตตัวล่างสูงกว่าปกติ ({vital.blood_pressure_dia} mmHg)")
    
    # ตรวจสอบอัตราการหายใจ
    if vital.respiratory_rate is not None:
        if vital.respiratory_rate < 12:
            abnormal_values.append(f"อัตราการหายใจต่ำกว่าปกติ ({vital.respiratory_rate} ครั้ง/นาที)")
        elif vital.respiratory_rate > 20:
            abnormal_values.append(f"อัตราการหายใจสูงกว่าปกติ ({vital.respiratory_rate} ครั้ง/นาที)")
    
    # ตรวจสอบความอิ่มตัวของออกซิเจนในเลือด
    if vital.oxygen_saturation is not None:
        if vital.oxygen_saturation < 95:
            abnormal_values.append(f"ความอิ่มตัวของออกซิเจนในเลือดต่ำกว่าปกติ ({vital.oxygen_saturation}%)")
    
    return abnormal_values



@app.route("/patient/<patient_id>")
def patient_vitals(patient_id):
    # ดึงข้อมูลผู้ป่วย
    patient_data = db.session.query(
        VitalSign.patient_name, 
        VitalSign.room_number,
        VitalSign.weight,
        VitalSign.height
    ).filter_by(patient_id=patient_id).first()
    
    if not patient_data:
        flash('ไม่พบข้อมูลผู้ป่วย', 'danger')
        return redirect(url_for('index'))
    
    # ดึงข้อมูล vital signs เรียงตามเวลา (ล่าสุดไปเก่าสุด) และกรองเฉพาะรายการที่มี vital signs จริง
    vitals = VitalSign.query.filter_by(patient_id=patient_id).filter(
        VitalSign.temperature != None,  # กรองเฉพาะรายการที่มีข้อมูล
        VitalSign.heart_rate != None,
        VitalSign.blood_pressure_sys != None
    ).order_by(desc(VitalSign.timestamp)).all()
    
    # ตรวจสอบค่าผิดปกติสำหรับการแสดงผลในหน้าสรุป
    abnormal_values = []
    has_abnormal_values = False
    
    if vitals:
        # ตรวจสอบค่าล่าสุด
        abnormal_values = check_abnormal_values(vitals[0])
        has_abnormal_values = len(abnormal_values) > 0
    
    return render_template(
        "patient.html", 
        patient_id=patient_id, 
        patient_name=patient_data.patient_name, 
        room_number=patient_data.room_number,
        weight=patient_data.weight,
        height=patient_data.height,
        vitals=vitals,
        abnormal_values=abnormal_values,
        has_abnormal_values=has_abnormal_values
    )

@app.route("/add_vital", methods=["GET", "POST"])
@app.route("/add_vital/<patient_id>", methods=["GET", "POST"])
def add_vital(patient_id=None):
    if request.method == "POST":
        try:
            # บันทึกเวลาเป็น UTC แต่บวกเวลา 7 ชั่วโมงเพื่อให้เป็นเวลาประเทศไทย
            thai_time = datetime.now() + timedelta(hours=7)
            
            new_vital = VitalSign(
                patient_id=request.form["patient_id"],
                patient_name=request.form["patient_name"],
                timestamp=thai_time,
                temperature=float(request.form["temperature"]),
                heart_rate=int(request.form["heart_rate"]),
                blood_pressure_sys=int(request.form["blood_pressure_sys"]),
                blood_pressure_dia=int(request.form["blood_pressure_dia"]),
                respiratory_rate=int(request.form["respiratory_rate"]),
                oxygen_saturation=int(request.form["oxygen_saturation"])
            )
            db.session.add(new_vital)
            db.session.commit()
            
            flash("บันทึกข้อมูล Vital Signs สำเร็จ", "success")
            return redirect(url_for("patient_vitals", patient_id=request.form["patient_id"]))
        except Exception as e:
            flash(f"เกิดข้อผิดพลาด: {str(e)}", "danger")
    
    # ดึงข้อมูลผู้ป่วยล่าสุดของแต่ละรหัสผู้ป่วยโดยไม่ซ้ำกัน
    from sqlalchemy import func

    
    # ใช้ subquery เพื่อดึงข้อมูลล่าสุดของแต่ละรหัสผู้ป่วย
    subquery = db.session.query(
        VitalSign.patient_id,
        func.max(VitalSign.timestamp).label('latest_timestamp')
    ).group_by(VitalSign.patient_id).subquery()
    
    patients = db.session.query(
        VitalSign.patient_id, 
        VitalSign.patient_name,
        VitalSign.room_number,
        VitalSign.weight,
        VitalSign.height
    ).join(
        subquery,
        db.and_(
            VitalSign.patient_id == subquery.c.patient_id,
            VitalSign.timestamp == subquery.c.latest_timestamp
        )
    ).all()
    
    # ถ้าไม่มีข้อมูลผู้ป่วย ให้เพิ่มข้อมูลตัวอย่าง
    if not patients:
        # สร้างข้อมูลตัวอย่างก่อนส่งไปหน้าฟอร์ม
        create_sample_data()
        
        # ดึงข้อมูลอีกครั้งหลังจากสร้างข้อมูลตัวอย่าง
        patients = db.session.query(
            VitalSign.patient_id, 
            VitalSign.patient_name,
            VitalSign.room_number,
            VitalSign.weight,
            VitalSign.height
        ).join(
            subquery,
            db.and_(
                VitalSign.patient_id == subquery.c.patient_id,
                VitalSign.timestamp == subquery.c.latest_timestamp
            )
        ).all()
    
    # ตัวแปรสำหรับเก็บ ID ของผู้ป่วยที่เลือก (ถ้ามี)
    selected_patient = None
    
    # ถ้ามีการระบุ patient_id
    if patient_id:
        # ค้นหาข้อมูลผู้ป่วยจากรายการ
        selected_patient = next((p for p in patients if p[0] == patient_id), None)
    
    return render_template("add_vital.html", patients=patients, selected_patient=selected_patient)

# ฟังก์ชันสำหรับดาวน์โหลดข้อมูลผู้ป่วยเป็นไฟล์ Excel
@app.route("/export/<patient_id>")
def export_patient_data(patient_id):
    # ดึงข้อมูลผู้ป่วย
    patient = db.session.query(VitalSign).filter_by(patient_id=patient_id).first()
    
    if not patient:
        flash('ไม่พบข้อมูลผู้ป่วย', 'danger')
        return redirect(url_for('index'))
    
    # ดึงข้อมูล vital signs เรียงตามเวลา (ล่าสุดไปเก่าสุด)
    vitals = VitalSign.query.filter_by(patient_id=patient_id).order_by(desc(VitalSign.timestamp)).all()
    
    # สร้าง DataFrame จากข้อมูล
    data = []
    for vital in vitals:
        # แปลงเวลาเป็นเวลาไทย
        thai_time = vital.timestamp + timedelta(hours=7) if vital.timestamp else None
        
        # ตรวจสอบว่ามีข้อมูล vital signs หรือไม่
        if vital.temperature is not None:
            data.append({
                'วันที่': thai_time.strftime("%d/%m/%Y") if thai_time else "",
                'เวลา': thai_time.strftime("%H:%M") if thai_time else "",
                'อุณหภูมิ (°C)': vital.temperature,
                'อัตราการเต้นของหัวใจ (ครั้ง/นาที)': vital.heart_rate,
                'ความดันโลหิตตัวบน (mmHg)': vital.blood_pressure_sys,
                'ความดันโลหิตตัวล่าง (mmHg)': vital.blood_pressure_dia,
                'อัตราการหายใจ (ครั้ง/นาที)': vital.respiratory_rate,
                'ออกซิเจนในเลือด (%)': vital.oxygen_saturation
            })
    
    if not data:
        flash('ไม่มีข้อมูล Vital Signs สำหรับผู้ป่วยนี้', 'warning')
        return redirect(url_for('patient_vitals', patient_id=patient_id))
    
    # สร้าง DataFrame
    df = pd.DataFrame(data)
    
    # สร้างไฟล์ Excel ในหน่วยความจำ
    output = io.BytesIO()
    
    # สร้าง ExcelWriter
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # เพิ่มข้อมูลลงในชีท
        df.to_excel(writer, sheet_name='Vital Signs', index=False)
        
        # ปรับแต่งรูปแบบของชีท
        workbook = writer.book
        worksheet = writer.sheets['Vital Signs']
        
        # ปรับความกว้างของคอลัมน์
        for idx, col in enumerate(df.columns):
            column_width = max(len(col) + 2, df[col].astype(str).map(len).max() + 2)
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(idx + 1)].width = column_width
        
        # ตกแต่งส่วนหัวตาราง
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # กำหนดรูปแบบของส่วนหัวตาราง
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # เพิ่มข้อมูลผู้ป่วยที่ด้านบนของชีท
        worksheet.insert_rows(1, 5)
        
        patient_info = VitalSign.query.filter_by(patient_id=patient_id).first()
        
        # เพิ่มหัวข้อรายงาน
        worksheet['A1'] = 'รายงานข้อมูล Vital Signs'
        worksheet['A1'].font = Font(bold=True, size=14)
        worksheet.merge_cells('A1:H1')
        worksheet['A1'].alignment = Alignment(horizontal='center')
        
        # เพิ่มข้อมูลผู้ป่วย
        worksheet['A2'] = 'รหัสผู้ป่วย:'
        worksheet['B2'] = patient_info.patient_id
        worksheet['A3'] = 'ชื่อ-นามสกุล:'
        worksheet['B3'] = patient_info.patient_name
        worksheet['A4'] = 'ห้อง:'
        worksheet['B4'] = patient_info.room_number or 'ไม่ระบุ'
        worksheet['C4'] = 'น้ำหนัก (กก.):'
        worksheet['D4'] = patient_info.weight or 'ไม่ระบุ'
        worksheet['E4'] = 'ส่วนสูง (ซม.):'
        worksheet['F4'] = patient_info.height or 'ไม่ระบุ'
        
        worksheet['A5'] = 'วันที่พิมพ์รายงาน:'
        thai_now = (datetime.now() + timedelta(hours=7)).strftime('%d/%m/%Y %H:%M')
        worksheet['B5'] = thai_now
        
        # ทำให้หัวข้อเป็นตัวหนา
        for row in range(2, 6):
            for col in ['A', 'C', 'E']:
                if worksheet[f'{col}{row}'].value:
                    worksheet[f'{col}{row}'].font = Font(bold=True)
    
    # ตั้งค่า pointer ไปที่จุดเริ่มต้นของไฟล์
    output.seek(0)
    
# ส่งไฟล์กลับไปให้ผู้ใช้
    return send_file(
        output,
        as_attachment=True,
        download_name=f"vital_signs_{patient_id}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.context_processor
def utility_processor():
    def format_datetime(dt):
        # แปลงเวลาให้เป็นเวลาประเทศไทย (UTC+7)
        thai_time = dt + timedelta(hours=7)
        return thai_time.strftime("%d/%m/%Y %H:%M")
    
    def format_date(dt):
        # แปลงเวลาให้เป็นเวลาประเทศไทย (UTC+7)
        thai_time = dt + timedelta(hours=7)
        return thai_time.strftime("%d/%m/%Y")
    
    def format_time(dt):
        # แปลงเวลาให้เป็นเวลาประเทศไทย (UTC+7)
        thai_time = dt + timedelta(hours=7)
        return thai_time.strftime("%H:%M")
    
    def get_vital_status(value, vital_type):
        # ถ้าค่าเป็น None ให้คืนค่าว่าง
        if value is None:
            return ""
            
        # กำหนดเกณฑ์สีสำหรับแต่ละประเภทของ vital sign
        if vital_type == "temperature":
            if value < 36.0:
                return "text-blue-700 font-bold"
            elif value > 38.0:
                return "text-red-700 font-bold"
            return "text-green-700"
        elif vital_type == "heart_rate":
            if value < 60:
                return "text-blue-700 font-bold"
            elif value > 100:
                return "text-red-700 font-bold"
            return "text-green-700"
        elif vital_type == "blood_pressure_sys":
            if value < 90:
                return "text-blue-700 font-bold"
            elif value > 140:
                return "text-red-700 font-bold"
            return "text-green-700"
        elif vital_type == "blood_pressure_dia":
            if value < 60:
                return "text-blue-700 font-bold"
            elif value > 90:
                return "text-red-700 font-bold"
            return "text-green-700"
        elif vital_type == "respiratory_rate":
            if value < 12:
                return "text-blue-700 font-bold"
            elif value > 20:
                return "text-red-700 font-bold"
            return "text-green-700"
        elif vital_type == "oxygen_saturation":
            if value < 95:
                return "text-red-700 font-bold"
            return "text-green-700"
        return ""
    
    return dict(
        format_datetime=format_datetime,
        format_date=format_date,
        format_time=format_time,
        get_vital_status=get_vital_status,
        datetime=datetime,
        timedelta=timedelta
    )

# ฟังก์ชันกล้อง #

def camera_stream():
    """ฟังก์ชันสำหรับเปิดกล้องและอ่านเฟรม"""
    global global_frame, camera_active
    
    # เปิดกล้อง (0 คือกล้องตัวแรก, สามารถเปลี่ยนเป็น 1, 2, ... สำหรับกล้องตัวอื่น)
    camera = cv2.VideoCapture(0)
    
    # ตรวจสอบว่าเปิดกล้องสำเร็จหรือไม่
    if not camera.isOpened():
        print("ไม่สามารถเปิดกล้องได้")
        return
    
    # ปรับความละเอียดของภาพ (ถ้าต้องการ)
    camera.set(cv2.CAP_PROP_FRAME_WIDTH, 640)
    camera.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)
    
    camera_active = True
    
    # วนลูปอ่านภาพจากกล้อง
    while camera_active:
        success, frame = camera.read()
        if not success:
            print("ไม่สามารถอ่านเฟรมจากกล้องได้")
            break
        
        # ตลับภาพถ้าคุณต้องการ
        # frame = cv2.flip(frame, 1)  # กลับด้านซ้าย-ขวา
        
        # แปลงเป็นรูปแบบที่เว็บแสดงผลได้
        _, buffer = cv2.imencode('.jpg', frame)
        encoded_frame = base64.b64encode(buffer).decode('utf-8')
        
        # อัปเดตเฟรมล่าสุด
        with lock:
            global_frame = encoded_frame
        
        # หน่วงเวลาเล็กน้อยเพื่อลดการใช้ CPU
        time.sleep(0.03)  # ประมาณ 30 FPS
    
    # เมื่อสิ้นสุดลูป ปิดกล้อง
    camera.release()
    print("ปิดกล้องแล้ว")

def gen_camera_frames():
    """ฟังก์ชันเพื่อส่งเฟรมของภาพเป็น multipart response"""
    global global_frame
    
    while True:
        # รอจนกว่าจะมีเฟรมล่าสุด
        with lock:
            if global_frame is None:
                continue
            frame_data = global_frame
        
        # ส่งข้อมูลรูปภาพในรูปแบบที่เหมาะสมสำหรับ multipart response
        yield (b'--frame\r\n'
               b'Content-Type: image/jpeg\r\n\r\n' + base64.b64decode(frame_data) + b'\r\n')
        
        # หน่วงเวลาเล็กน้อย
        time.sleep(0.03)

# เส้นทางสำหรับเปิด/ปิดกล้อง
@app.route('/start_camera')
def start_camera():
    global camera_thread, camera_active, global_frame
    
    if camera_thread is None or not camera_thread.is_alive():
        global_frame = None
        camera_active = True
        camera_thread = threading.Thread(target=camera_stream)
        camera_thread.daemon = True
        camera_thread.start()
        return "กำลังเปิดกล้อง..."
    
    return "กล้องเปิดอยู่แล้ว"

@app.route('/stop_camera')
def stop_camera():
    global camera_active, camera_thread
    
    camera_active = False
    if camera_thread is not None and camera_thread.is_alive():
        camera_thread.join(timeout=1.0)
    
    return "ปิดกล้องแล้ว"

# เส้นทางสำหรับสตรีมวิดีโอ
@app.route('/video_feed')
def video_feed():
    return Response(gen_camera_frames(),
                    mimetype='multipart/x-mixed-replace; boundary=frame')

# เส้นทางสำหรับหน้ากล้อง
@app.route('/camera')
def camera():
    return render_template('camera.html')

# เส้นทางสำหรับบันทึกภาพถ่าย
@app.route('/save_image', methods=['POST'])
def save_image():
    try:
        # รับข้อมูลรูปภาพจากคำขอ
        data = request.get_json()
        if not data or 'image_data' not in data:
            return jsonify({'success': False, 'error': 'ไม่พบข้อมูลรูปภาพ'}), 400
        
        # แยกส่วนหัวของ Data URL ออกจากข้อมูลรูปภาพ
        image_data = data['image_data']
        if 'base64,' in image_data:
            # ถ้าเป็น Data URL แบบเต็ม (เช่น data:image/png;base64,ABC123...)
            image_data = image_data.split('base64,')[1]
        
        # สร้างชื่อไฟล์ด้วยวันที่และเวลาปัจจุบัน
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # เพิ่ม patient_id ถ้ามี
        patient_id = data.get('patient_id', '')
        if patient_id:
            filename = f"patient_{patient_id}_{timestamp}.png"
        else:
            filename = f"image_{timestamp}.png"
        
        # สร้างเส้นทางไฟล์เต็ม
        filepath = os.path.join(UPLOAD_FOLDER, filename)
        
        # แปลงข้อมูล base64 เป็นไฟล์ภาพและบันทึก
        with open(filepath, "wb") as f:
            f.write(base64.b64decode(image_data))
        
        # คืนค่า URL ของรูปภาพที่บันทึกแล้ว
        image_url = f"/static/uploads/{filename}"
        return jsonify({
            'success': True, 
            'message': 'บันทึกรูปภาพสำเร็จ',
            'filename': filename,
            'url': image_url
        })
    
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/robot_control')
def robot_control():
    return render_template('robot_control.html')

# API สำหรับควบคุมหุ่นยนต์
@app.route('/control', methods=['POST'])
def control_robot():
    data = request.json
    linear_x = data.get('linear_x', 0.0)
    angular_z = data.get('angular_z', 0.0)
    
    robot_controller.publish_velocity(linear_x, angular_z)
    return jsonify({'status': 'success', 'linear_x': linear_x, 'angular_z': angular_z})

@socketio.on('nav_goal')
def handle_nav_goal(data):
    """รับข้อมูลเป้าหมายการนำทางจาก Socket.IO และส่งไปยัง robot_controller"""
    print(f"Received navigation goal: {data}")
    if robot_controller:
        try:
            x = float(data.get('x', 0.0))
            y = float(data.get('y', 0.0))
            theta = float(data.get('theta', 0.0))
            robot_controller.send_nav_goal(x, y, theta)
        except Exception as e:
            print(f"Error sending navigation goal: {e}")
    else:
        socketio.emit('debug_message', {'message': 'Robot controller not available'})

# ฟังก์ชันสำหรับเตรียมฐานข้อมูลและสร้างข้อมูลตัวอย่าง
def initialize_database():
    with app.app_context():
        # สร้างข้อมูลตัวอย่างทันที
        create_sample_data()
        print("สร้างข้อมูลตัวอย่างสำเร็จ! จำนวนข้อมูล:", VitalSign.query.count())

@socketio.on('request_real_map')
def handle_request_real_map():
    """รับคำขอข้อมูลแผนที่จริงโดยตรง"""
    print("Client requested real map")
    if robot_controller:
        robot_controller.handle_request_real_map()
    else:
        socketio.emit('debug_message', {'message': 'Robot controller not available'})

# เพิ่มการ import sqlalchemy exceptions ที่ต้นไฟล์ main.py
from sqlalchemy import exc as sqlalchemy_exc

# จัดการกับ HTTP errors
@app.errorhandler(404)
def page_not_found(e):
    flash("ไม่พบหน้าที่คุณต้องการ", "warning")
    return redirect(url_for('index'))

@app.errorhandler(500)
def internal_server_error(e):
    flash("เกิดข้อผิดพลาดในระบบ กรุณาลองใหม่อีกครั้ง", "danger")
    return redirect(url_for('index'))

# จัดการกับข้อผิดพลาดอื่นๆ ทั้งหมด
@app.errorhandler(Exception)
def handle_exception(e):
    # แสดงข้อความข้อผิดพลาดในคอนโซล (ช่วยในการแก้ไขปัญหา)
    print(f"เกิดข้อผิดพลาด: {str(e)}")
    
    # แสดงข้อความแจ้งเตือนผู้ใช้
    flash("เกิดข้อผิดพลาดในระบบ กรุณาลองใหม่อีกครั้ง", "danger")
    
    # เด้งกลับไปหน้าหลัก
    return redirect(url_for('index'))

# ถ้าคุณต้องการจัดการเฉพาะบาง errors
@app.errorhandler(TypeError)
def handle_type_error(e):
    flash("เกิดข้อผิดพลาดในการประมวลผลข้อมูล กรุณาตรวจสอบและลองใหม่อีกครั้ง", "danger")
    return redirect(url_for('index'))

@app.errorhandler(ValueError)
def handle_value_error(e):
    flash("ข้อมูลไม่ถูกต้อง กรุณาตรวจสอบและลองใหม่อีกครั้ง", "danger")
    return redirect(url_for('index'))

# หากคุณต้องการจัดการกับ SQLAlchemy errors
@app.errorhandler(sqlalchemy_exc.SQLAlchemyError)
def handle_db_error(e):
    db.session.rollback()  # ยกเลิกการทำธุรกรรมที่ไม่สมบูรณ์
    flash("เกิดข้อผิดพลาดในการเข้าถึงฐานข้อมูล กรุณาลองใหม่อีกครั้ง", "danger")
    return redirect(url_for('index'))

@app.route('/webrtc_stream')
def webrtc_stream():
    return render_template('webrtc_stream.html')


if __name__ == '__main__':
    try:
        socketio.run(app, debug=True, host='0.0.0.0', port=5000, allow_unsafe_werkzeug=True)
    finally:
        robot_controller.shutdown()